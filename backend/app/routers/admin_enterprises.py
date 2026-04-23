from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from app.database import get_db
from app.models.enterprise import Enterprise
from app.dependencies import get_current_user

router = APIRouter(prefix="/api/admin/enterprises", tags=["enterprises"])


class EnterpriseCreate(BaseModel):
    customer_name: str
    province: str
    city: str
    industry: str
    company_intro: str = ""
    yonyou_content: str = ""


class EnterpriseUpdate(BaseModel):
    customer_name: Optional[str] = None
    province: Optional[str] = None
    city: Optional[str] = None
    industry: Optional[str] = None
    company_intro: Optional[str] = None
    yonyou_content: Optional[str] = None


@router.get("")
def list_enterprises(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    industry: Optional[str] = None,
    province: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    query = db.query(Enterprise)
    if industry:
        query = query.filter(Enterprise.industry == industry)
    if province:
        query = query.filter(Enterprise.province == province)
    if keyword:
        query = query.filter(Enterprise.customer_name.contains(keyword))

    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return {
        "items": [
            {
                "id": e.id,
                "customer_name": e.customer_name,
                "province": e.province,
                "city": e.city,
                "industry": e.industry,
                "company_intro": e.company_intro or "",
                "yonyou_content": e.yonyou_content or "",
            }
            for e in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("")
def create_enterprise(data: EnterpriseCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    enterprise = Enterprise(**data.model_dump())
    db.add(enterprise)
    db.commit()
    db.refresh(enterprise)
    return {"id": enterprise.id, "message": "创建成功"}


@router.put("/{enterprise_id}")
def update_enterprise(enterprise_id: int, data: EnterpriseUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    enterprise = db.query(Enterprise).filter(Enterprise.id == enterprise_id).first()
    if not enterprise:
        raise HTTPException(status_code=404, detail="企业不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(enterprise, key, value)

    db.commit()
    return {"message": "更新成功"}


@router.delete("/{enterprise_id}")
def delete_enterprise(enterprise_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    enterprise = db.query(Enterprise).filter(Enterprise.id == enterprise_id).first()
    if not enterprise:
        raise HTTPException(status_code=404, detail="企业不存在")

    db.delete(enterprise)
    db.commit()
    return {"message": "删除成功"}


@router.post("/import")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    import tempfile
    import os
    from openpyxl import load_workbook

    COLUMN_MAP = {
        "客户名称": "customer_name",
        "客户所在省": "province",
        "客户所在市": "city",
        "标准行业": "industry",
        "企业简介": "company_intro",
        "用友建设内容": "yonyou_content",
    }

    # 文件类型校验
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls 文件")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")

    # 保存上传文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tmp.write(contents)
        tmp.close()

        wb = load_workbook(tmp.name, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_index = {h: i for i, h in enumerate(headers) if h}

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            kwargs = {}
            for cn_name, field in COLUMN_MAP.items():
                idx = header_index.get(cn_name)
                if idx is not None:
                    kwargs[field] = row[idx] if row[idx] is not None else ""
            if kwargs.get("customer_name"):
                db.add(Enterprise(**kwargs))
                count += 1

        wb.close()
        db.commit()
        return {"message": f"成功导入 {count} 条记录"}
    finally:
        os.unlink(tmp.name)
