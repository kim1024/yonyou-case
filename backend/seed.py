"""seed.py - 从 Excel 导入企业数据，自动创建管理员账号，初始化基础数据。"""

import logging
import sys
from pathlib import Path

# 将 backend/ 加入 sys.path，确保 `app.*` 可被导入
BACKEND_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BACKEND_DIR))

from app.logging_config import setup_logging

setup_logging()

_logger = logging.getLogger(__name__)

import yaml
from openpyxl import load_workbook

from app.database import Base, engine, SessionLocal
from app.models.enterprise import Enterprise
from app.models.admin import AdminUser
from app.models.major import Major, Industry, MajorIndustry, Region, Hour
from app.models.llm_config import LLMConfig
from app.models.prompt_template import PromptTemplate
from app.models.prompt_version import PromptVersion
from app.models.province_city import Province, City
from app.services.auth_service import get_password_hash

# ---------- 路径 ----------
PROJECT_ROOT = BACKEND_DIR.parent
EXCEL_PATH = PROJECT_ROOT / "old" / "data" / "data.xlsx"
CONFIG_PATH = PROJECT_ROOT / "config.yaml"

# Excel 中文列名 → 模型字段映射
COLUMN_MAP = {
    "客户名称": "customer_name",
    "客户所在省": "province",
    "客户所在市": "city",
    "标准行业": "industry",
    "企业简介": "company_intro",
    "用友建设内容": "yonyou_content",
}


def _load_config() -> dict:
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}


def seed_enterprises(db):
    """从 Excel 导入企业数据（仅在表为空时执行）。"""
    if db.query(Enterprise).first() is not None:
        _logger.info("enterprises table already has data, skipping import.")
        return

    if not EXCEL_PATH.exists():
        _logger.error("Excel file not found: %s", EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {h: i for i, h in enumerate(headers) if h}

    enterprises = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        kwargs = {}
        for cn_name, field in COLUMN_MAP.items():
            idx = header_index.get(cn_name)
            if idx is not None:
                kwargs[field] = row[idx] if row[idx] is not None else ""
        # 至少需要 customer_name
        if kwargs.get("customer_name"):
            enterprises.append(Enterprise(**kwargs))

    wb.close()

    if enterprises:
        db.add_all(enterprises)
        db.commit()
        _logger.info("Successfully imported %d enterprise records.", len(enterprises))
    else:
        _logger.warning("No valid data parsed from Excel.")


def seed_admin(db):
    """如果 admin_users 表为空，则从 config.yaml 创建管理员账号。"""
    if db.query(AdminUser).first() is not None:
        _logger.info("admin_users table already has data, skipping creation.")
        return

    config = _load_config()
    admin_cfg = config.get("admin", {})
    username = admin_cfg.get("username", "admin")
    password = admin_cfg.get("password", "changeme")

    user = AdminUser(
        username=username,
        password_hash=get_password_hash(password),
    )
    db.add(user)
    db.commit()
    _logger.info("Created admin user: %s", username)


def seed_provinces_cities(db):
    """从 enterprises 表中提取省市数据，创建 Province 和 City 记录。"""
    # 检查是否已存在数据
    if db.query(Province).first() is not None:
        _logger.info("provinces table already has data, skipping.")
        return

    # 从 enterprises 表提取 distinct (province, city) 组合
    results = db.query(Enterprise.province, Enterprise.city).distinct().all()

    province_map = {}
    for province_name, city_name in results:
        if not province_name or province_name.strip() == '':
            continue

        # 创建或获取省份
        if province_name not in province_map:
            province = Province(name=province_name)
            db.add(province)
            db.flush()  # 获取 id
            province_map[province_name] = province

        # 创建城市
        if city_name and city_name.strip() != '':
            city = City(name=city_name, province_id=province_map[province_name].id)
            db.add(city)

    db.commit()
    _logger.info("Created %d province and %d city records.", len(province_map),
                 db.query(City).count())


def seed_database():
    """主入口：建表 + 导入数据。"""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        seed_enterprises(db)
        seed_admin(db)
        seed_majors(db)
        seed_industries(db)
        seed_regions(db)
        seed_hours(db)
        seed_llm_config(db)
        seed_prompt_templates(db)
        seed_provinces_cities(db)
    finally:
        db.close()
    _logger.info("Seed completed.")


# ---------- 基础数据 ----------

def seed_majors(db):
    """初始化专业数据。数据不匹配时自动修正。"""
    expected = [
        {"name": "大数据", "description": "聚焦数据采集、存储、分析与可视化等核心技术", "icon": "BarChart3", "sort_order": 1},
        {"name": "人工智能", "description": "聚焦机器学习、深度学习与智能应用开发", "icon": "Brain", "sort_order": 2},
        {"name": "工业互联网", "description": "聚焦工业物联网、智能制造与数字化转型", "icon": "Cpu", "sort_order": 3},
    ]
    expected_names = {m["name"] for m in expected}
    existing_names = {m.name for m in db.query(Major).all()}

    if existing_names == expected_names:
        _logger.info("majors data already correct, skipping.")
        return

    # 数据不匹配，清除旧数据重建
    if existing_names:
        _logger.info("Majors data mismatch (got: %s), rebuilding.", existing_names)
        db.query(MajorIndustry).delete()  # 先删关联表
        db.query(Major).delete()
        db.commit()

    for data in expected:
        db.add(Major(**data))
    db.commit()
    _logger.info("Created %d major records.", len(expected))


def seed_industries(db):
    """从 enterprises 表提取行业去重，创建 Industry 记录及 MajorIndustry 关联。"""
    existing_industries = db.query(Industry).first() is not None
    existing_associations = db.query(MajorIndustry).first() is not None

    if not existing_industries:
        # 行业表为空，创建行业和关联
        rows = db.query(Enterprise.industry).distinct().all()
        industry_names = sorted({r[0] for r in rows if r[0]})

        if not industry_names:
            _logger.warning("No industries found in enterprises table.")
            return

        majors = db.query(Major).order_by(Major.sort_order).all()
        major_count = len(majors)

        industry_id_map: dict[str, int] = {}
        for idx, name in enumerate(industry_names):
            industry = Industry(name=name, sort_order=idx + 1)
            db.add(industry)
            db.flush()
            industry_id_map[name] = industry.id

        associations = []
        for idx, name in enumerate(industry_names):
            if major_count > 0:
                major_id = majors[idx % major_count].id
                associations.append(
                    MajorIndustry(major_id=major_id, industry_id=industry_id_map[name])
                )
        if associations:
            db.add_all(associations)

        db.commit()
        _logger.info(
            "Created %d industry records and %d MajorIndustry associations.",
            len(industry_names), len(associations),
        )
        return

    # 行业表已有数据，检查 MajorIndustry 关联是否完整
    industry_count = db.query(Industry).count()
    association_count = db.query(MajorIndustry).count()

    if existing_associations and association_count > 0:
        _logger.info("industries and MajorIndustry associations already exist, skipping.")
        return

    # 行业有数据但关联缺失，重建关联
    _logger.info("MajorIndustry associations missing (industries: %d), rebuilding.", industry_count)
    majors = db.query(Major).order_by(Major.sort_order).all()
    major_count = len(majors)
    if major_count == 0:
        _logger.warning("No majors found, cannot rebuild MajorIndustry associations.")
        return

    industries = db.query(Industry).order_by(Industry.sort_order).all()
    associations = []
    for idx, industry in enumerate(industries):
        major_id = majors[idx % major_count].id
        associations.append(
            MajorIndustry(major_id=major_id, industry_id=industry.id)
        )

    if associations:
        db.add_all(associations)
        db.commit()
        _logger.info("Rebuilt %d MajorIndustry associations.", len(associations))


def seed_regions(db):
    """从 enterprises 表提取省份去重，创建 Region 记录。"""
    if db.query(Region).first() is not None:
        _logger.info("regions table already has data, skipping.")
        return

    rows = db.query(Enterprise.province).distinct().all()
    region_names = sorted({r[0] for r in rows if r[0]})

    if not region_names:
        _logger.warning("No regions found in enterprises table.")
        return

    for idx, name in enumerate(region_names):
        db.add(Region(name=name, sort_order=idx + 1))

    db.commit()
    _logger.info("Created %d region records.", len(region_names))


def seed_hours(db):
    """初始化课时数据（仅在表为空时执行）。"""
    if db.query(Hour).first() is not None:
        _logger.info("hours table already has data, skipping.")
        return

    hours_data = [
        {"value": 8,  "label": "8课时（1天）",  "sort_order": 1},
        {"value": 16, "label": "16课时（2天）", "sort_order": 2},
        {"value": 24, "label": "24课时（3天）", "sort_order": 3},
        {"value": 32, "label": "32课时（4天）", "sort_order": 4},
    ]
    for data in hours_data:
        db.add(Hour(**data))
    db.commit()
    _logger.info("Created %d hour records.", len(hours_data))


def seed_llm_config(db):
    """从 config.yaml 读取 llm 配置，首次运行时写入数据库作为默认配置。"""
    if db.query(LLMConfig).first() is not None:
        _logger.info("llm_configs table already has data, skipping.")
        return

    config = _load_config()
    llm_cfg = config.get("llm", {})

    api_key = llm_cfg.get("api_key", "")
    if not api_key or api_key == "sk-xxx":
        _logger.warning("No valid LLM API key found in config.yaml, skipping llm_config seed.")
        return

    llm_config = LLMConfig(
        name="默认模型",
        api_base_url=llm_cfg.get("api_base_url", "https://api.openai.com/v1"),
        api_key=api_key,
        model=llm_cfg.get("model", "gpt-4o"),
        temperature=llm_cfg.get("temperature", 0.7),
        max_tokens=llm_cfg.get("max_tokens", 2000),
        timeout=llm_cfg.get("timeout", 60),
        is_active=True,
    )
    db.add(llm_config)
    db.commit()
    _logger.info("Created default LLM config from config.yaml (model: %s).", llm_config.model)


def seed_prompt_templates(db):
    """创建默认提示词模板（仅在表为空时执行）。"""
    if db.query(PromptTemplate).first() is not None:
        _logger.info("prompt_templates table already has data, skipping.")
        return

    # 主提示词模板（AI 生成用，要求 JSON 输出）
    main_prompt_content = """请根据以下信息，生成一份产业案例教学课程设计方案。

专业方向：{major}
行业：{industry}
企业：{enterprise_name}
地区：{region}
课时：{hour}课时

<企业简介>
{company_intro}
</企业简介>
<用友建设内容>
{yonyou_content}
</用友建设内容>

课时分配参考：
- 模块一（行业背景与需求分析）：{hour_block1}课时
- 模块二（技术基础与工具介绍）：{hour_block2}课时
- 模块三（案例实战与项目实施）：{hour_block3}课时
- 模块四（总结与拓展）：{hour_block4}课时

请严格按照以下 JSON 结构输出（仅输出 JSON，不要输出任何其他内容）：

{{
  "title": "{enterprise_name}案例教学课程方案",
  "subtitle": "{enterprise_name}",
  "introduction": "（请根据实际信息丰富此段介绍，不少于100字。说明本教学案例基于{enterprise_name}公司的真实业务场景，结合{industry}专业技术，设计了{hour}课时教学方案。需要强调的动态内容（如企业名、行业名、专业名、课时数等）请用 HTML 标签 <b class=\\"highlight\\">内容</b> 包裹，使其加粗并使用特殊颜色显示。）",
  "modules": [
    {{
      "name": "模块一：行业背景与需求分析",
      "hours": {hour_block1},
      "items": [
        "{industry}行业现状与发展趋势",
        "{enterprise_name}业务模式与技术需求分析",
        "数字化转型痛点与机遇"
      ]
    }},
    {{
      "name": "模块二：技术基础与工具介绍",
      "hours": {hour_block2},
      "items": [
        "{major}核心技术原理与架构",
        "用友产品体系与解决方案概览",
        "开发环境搭建与工具链配置"
      ]
    }},
    {{
      "name": "模块三：案例实战与项目实施",
      "hours": {hour_block3},
      "items": [
        "{enterprise_name}真实业务场景解析",
        "基于用友平台的功能开发与集成",
        "项目方案设计、实施与优化"
      ]
    }},
    {{
      "name": "模块四：总结与拓展",
      "hours": {hour_block4},
      "items": [
        "项目成果展示与答辩",
        "{industry}领域最佳实践总结",
        "职业发展路径与学习资源推荐"
      ]
    }}
  ],
  "positions": [
    {{
      "title": "{major}工程师",
      "description": [
        "负责{industry}领域的数据/系统开发",
        "参与企业数字化转型项目",
        "熟练使用用友产品体系"
      ]
    }},
    {{
      "title": "{industry}解决方案架构师",
      "description": [
        "设计行业数字化解决方案",
        "对接客户需求与技术实现"
      ]
    }},
    {{
      "title": "项目实施顾问",
      "description": [
        "负责用友产品在企业的落地实施",
        "提供客户培训与技术支持"
      ]
    }},
    {{
      "title": "业务分析师",
      "description": [
        "分析{industry}业务流程与需求",
        "设计数字化优化方案"
      ]
    }},
    {{
      "title": "技术项目经理",
      "description": [
        "管理{industry}领域IT项目",
        "协调团队与客户资源"
      ]
    }},
    {{
      "title": "数字化运营专员",
      "description": [
        "负责{industry}领域数字化运营与持续优化",
        "监控系统运行指标，推动业务流程改进"
      ]
    }}
  ],
  "deliverables": [
    "PPT课件",
    "教学视频",
    "实验指导书",
    "数据集",
    "代码包",
    "实操环境配置文档"
  ],
  "notes": "以上内容由 AI 生成，请结合实际教学需求进行调整。"
}}

重要提示：
1. 仅输出上述 JSON 对象，不要输出其他任何内容。
2. JSON 中所有字段均为必填项。
3. introduction 字段不少于100字。
4. modules 数组必须包含4个模块，每个模块的 items 不少于3条。
5. positions 数组必须包含6个岗位，岗位和描述需结合 {industry} 领域与 {major} 专业。
6. introduction 中需要强调的动态内容（企业名、行业名、专业名、课时数等），请用 HTML 标签包裹：<b class="highlight">xxx</b>，使其加粗并使用特殊颜色显示。
7. 报价信息不需要生成，由系统另行计算。"""

    # 创建模板
    template = PromptTemplate(
        name="课程方案生成模板",
        description="用于生成产业案例教学课程设计方案的提示词模板",
        scene="课程方案生成",
        is_active=True,
    )
    db.add(template)
    db.flush()

    # 创建版本 1：主提示词模板
    version = PromptVersion(
        template_id=template.id,
        version_number=1,
        content=main_prompt_content,
        variables='["major", "industry", "enterprise_name", "region", "hour", "company_intro", "yonyou_content", "hour_block1", "hour_block2", "hour_block3", "hour_block4"]',
        remark="初始版本：主提示词模板",
        created_by="system",
    )
    db.add(version)
    db.flush()

    # 设置当前版本
    template.current_version_id = version.id
    db.commit()
    _logger.info("Created default prompt template (scene: %s, version: 1).", template.scene)


if __name__ == "__main__":
    seed_database()
